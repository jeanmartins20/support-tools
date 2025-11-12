# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import pyodbc
from typing import Optional
import socket
import re
import configparser 
import os 
import threading
import queue
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter 
from openpyxl.styles import Font

CONFIG_FILE = 'apuracao_webphone.ini'

class WebPhoneReportTool(tk.Tk):
    """
    Ferramenta de apuração de dados WebPhone.
    Conecta no banco de leitura V2 e executa 4 queries
    específicas, exibindo os resultados em abas.
    """
    
    # --- QUERIES (Mesmas de antes) ---
    QUERY_PLANO = """
    select nome, valor
    from funcionalidadeplano
    where funcionalidadeid IN (4,18)
    and situacaoregistro = 1
    and trial = 0
    """

    QUERY_BASE_CLIENTE = """
    select  distinct '''' + e.id1 as 'CNPJ',
    e.id as 'SpotterId', 
    e.razaosocial,
    fp.nome as 'Plano',
    REPLACE(REPLACE(CONVERT(varchar(30), fp.valor, 1),',',''),'.',',') as Mensalidade
    from funcionalidadeplanoempresa fe
    join funcionalidadeplano fp on fp.id = fe.FuncionalidadePlanoId
    join EmpresaCliente e on e.id = fe.EmpresaClienteId
    where fe.funcionalidadeplanoid in (select id from funcionalidadeplano where funcionalidadeid IN (4,18)
    and situacaoregistro = 1
    and trial = 0)
    and fe.situacao = 1
    """
    
    QUERY_DETALHAMENTO = """
    select FORMAT(DATEADD(DAY,1,EOMONTH(getdate(),-2)), 'dd/MM/yyyy') as Data,
    ec.id as SpotterId, ec.RazaoSocial, '''' + ec.id1 as CNPJ,
    case when hl.TipoTelefone = 1 then 'Fixo' else 'Móvel' end TipoTelefone,
    f.Nome as Plano, 
    case 
        when hl.TipoTelefone = 1 
        then CAST(REPLACE(fpt.PrecoFixo,'.',',') AS varchar)
        else CAST(REPLACE(fpt.PrecoMovel,'.',',') AS varchar)
    end PrecoSegundo, 
    sum(hl.SegundosTarifaveis) as SegundosTarifaveis,
    CAST(REPLACE(sum(hl.Custo),'.',',') AS varchar) as Cobranca
    from HistoricoLigacao HL
    join EmpresaCliente EC on EC.Id = HL.EmpresaClienteId
    join FuncionalidadePlanoEmpresa fp on fp.empresaclienteid = ec.id AND fp.Situacao=1
    join funcionalidadeplano f on f.id = fp.FuncionalidadePlanoId AND f.SituacaoRegistro=1 AND f.Trial=0
    join FuncionalidadePlanoTarifa fpt on fpt.FuncionalidadePlanoId = f.id AND CONVERT(INT, hl.DDIDestinoTel) = fpt.Pais
    where   fp.FuncionalidadeId in (4,18)
        and HL.SegundosTarifaveis > 0
        and HL.ParceiroWebphone > 0
        and HL.CreatedAt >= DATEADD(DAY,1,EOMONTH(getdate(),-2))
        and HL.CreatedAt < DATEADD(DAY,1,EOMONTH(getdate(),-1))
    group by ec.id, ec.RazaoSocial, ec.id1, hl.TipoTelefone, fpt.PrecoFixo, fpt.PrecoMovel, f.Nome
    order by 1
    """

    QUERY_CREDITO = """
    select ec.id as SpotterId, '''' + ec.id1 as CNPJ, ec.razaosocial, 
    REPLACE(CAST(fp.credito AS DECIMAL(29,2)) ,'.',',') as creditowebphone
    from EmpresaCliente ec
    join FuncionalidadePlanoEmpresa fp on fp.empresaclienteid = ec.id and fp.situacao =1
    join funcionalidadeplano f on f.id = fp.FuncionalidadePlanoId and f.trial =0 and f.SituacaoRegistro=1
    where fp.credito > 0
    and fp.FuncionalidadeId in (4,18)
    order by ec.razaosocial
    """

    def __init__(self):
        super().__init__()
        self.title("Ferramenta de Apuração WebPhone")
        self.geometry("800x600")

        # --- Conexão ---
        self.conn: Optional[pyodbc.Connection] = None
        self.cursor: Optional[pyodbc.Cursor] = None
        
        # --- Widgets Conexão ---
        self.server_entry = None
        self.auth_type_combo = None
        self.db_entry = None
        self.logon_label = None
        self.logon_entry = None
        self.pass_label = None
        self.pass_entry = None
        self.status_label = None
        
        # --- Widgets Resultados ---
        self.load_data_button = None
        self.export_button = None
        self.report_status_label = None
        self.results_notebook = None
        self.tree_plano = None
        self.tree_base = None
        self.tree_detalhe = None
        self.tree_credito = None
        
        # --- Threading & Config ---
        self.query_queue = queue.Queue()
        self.config = configparser.ConfigParser()
        
        self._create_widgets()
        self._on_auth_type_change()
        self._load_config()
        self._monitor_queue()

    def _create_widgets(self):
        """Cria os componentes da interface (abas, campos, botões)."""
        
        # Notebook Principal (Conexão | Resultados)
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(pady=10, padx=10, fill="both", expand=True)

        # Lista de servidores (padrão do WA.py)
        server_list = [
            "spotterv1prd.database.windows.net",
            "spotterprdv2-leitura.database.windows.net",
            "spotterprdv2.database.windows.net",
            "queue-observer-prd.database.windows.net",
            "globalserviceprd.database.windows.net",
            "spotterprd-v2-leitura.database.windows.net"
        ]

        # --- ABA 1: Conexão ---
        self.conn_frame = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(self.conn_frame, text="Conexão")

        ttk.Label(self.conn_frame, text="Configurações de Conexão (Leitura V2)", 
                  font=("Arial", 14, "bold")).pack(pady=10)
        conn_grid = ttk.Frame(self.conn_frame)
        conn_grid.pack(pady=5, padx=5)

        ttk.Label(conn_grid, text="Tipo de servidor:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.server_type_combo = ttk.Combobox(conn_grid, width=38, state="readonly", 
                                              values=["Mecanismo de Banco de Dados"])
        self.server_type_combo.grid(row=0, column=1, padx=5, pady=5)
        self.server_type_combo.current(0)

        ttk.Label(conn_grid, text="Nome do servidor:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        self.server_entry = ttk.Combobox(conn_grid, width=38, values=server_list)
        self.server_entry.grid(row=1, column=1, padx=5, pady=5)
        self.server_entry.insert(0, "") # <-- CORRIGIDO

        ttk.Label(conn_grid, text="Autenticação:").grid(row=2, column=0, sticky="w", padx=5, pady=5)
        self.auth_type_combo = ttk.Combobox(conn_grid, width=38, state="readonly",
                                            values=["Autenticação do SQL Server", "Autenticação do Windows"])
        self.auth_type_combo.grid(row=2, column=1, padx=5, pady=5)
        self.auth_type_combo.current(0) 
        self.auth_type_combo.bind("<<ComboboxSelected>>", self._on_auth_type_change)

        ttk.Label(conn_grid, text="Banco de Dados:").grid(row=3, column=0, sticky="w", padx=5, pady=5)
        self.db_entry = ttk.Entry(conn_grid, width=40)
        self.db_entry.grid(row=3, column=1, padx=5, pady=5)
        self.db_entry.insert(0, "") # <-- CORRIGIDO

        self.logon_label = ttk.Label(conn_grid, text="Logon:")
        self.logon_label.grid(row=4, column=0, sticky="w", padx=5, pady=5)
        self.logon_entry = ttk.Entry(conn_grid, width=40)
        self.logon_entry.grid(row=4, column=1, padx=5, pady=5)
        self.logon_entry.insert(0, "") # <-- CORRIGIDO

        self.pass_label = ttk.Label(conn_grid, text="Senha:")
        self.pass_label.grid(row=5, column=0, sticky="w", padx=5, pady=5)
        self.pass_entry = ttk.Entry(conn_grid, width=40, show="*")
        self.pass_entry.grid(row=5, column=1, padx=5, pady=5)
        
        self.connect_button = ttk.Button(self.conn_frame, text="Conectar", command=self.connect_db)
        self.connect_button.pack(pady=20)

        self.status_label = ttk.Label(self.conn_frame, text="Status: Desconectado", 
                                      font=("Arial", 10), foreground="red")
        self.status_label.pack(pady=5)

        # --- ABA 2: Resultados ---
        self.report_frame = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(self.report_frame, text="Apuração WebPhone")
        
        # Frame de botões
        button_frame = ttk.Frame(self.report_frame)
        button_frame.pack(fill="x", pady=5)
        
        self.load_data_button = ttk.Button(button_frame, text="Carregar Dados", 
                                           command=self.start_load_data_thread)
        self.load_data_button.pack(side="left", padx=5)

        # --- BOTÃO MODIFICADO ---
        self.export_button = ttk.Button(button_frame, text="Exportar para Excel (.xlsx)",
                                        command=self.export_all_to_excel, state="disabled")
        self.export_button.pack(side="left", padx=5)

        self.report_status_label = ttk.Label(self.report_frame, text="Aguardando carga de dados...",
                                             font=("Arial", 10), foreground="blue")
        self.report_status_label.pack(anchor="w", fill="x", pady=(0, 5))

        # Notebook Aninhado (para as 4 tabelas)
        self.results_notebook = ttk.Notebook(self.report_frame)
        self.results_notebook.pack(fill="both", expand=True, pady=5)

        # Tabela 1: Planos
        self.tree_plano = self._create_tree_tab(self.results_notebook, "1. Atualização de Plano")
        # Tabela 2: Base Cliente
        self.tree_base = self._create_tree_tab(self.results_notebook, "2. Base de Cliente")
        # Tabela 3: Detalhamento
        self.tree_detalhe = self._create_tree_tab(self.results_notebook, "3. Detalhamento")
        # Tabela 4: Crédito
        self.tree_credito = self._create_tree_tab(self.results_notebook, "4. Base de Crédito")

        # Desabilita a aba de resultados por padrão
        self.notebook.tab(1, state="disabled")

    def _create_tree_tab(self, parent_notebook: ttk.Notebook, tab_name: str) -> ttk.Treeview:
        """Função auxiliar para criar uma aba com um Treeview."""
        frame = ttk.Frame(parent_notebook, padding="5")
        parent_notebook.add(frame, text=tab_name)
        
        tree_frame = ttk.Frame(frame)
        tree_frame.pack(fill="both", expand=True)
        
        scroll_y = ttk.Scrollbar(tree_frame, orient="vertical")
        scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal")
        
        tree = ttk.Treeview(tree_frame, 
                            yscrollcommand=scroll_y.set,
                            xscrollcommand=scroll_x.set,
                            height=15)
        
        scroll_y.config(command=tree.yview)
        scroll_x.config(command=tree.xview)
        scroll_y.pack(side="right", fill="y")
        scroll_x.pack(side="bottom", fill="x")
        tree.pack(fill="both", expand=True)
        
        return tree

    def _on_auth_type_change(self, event=None):
        """Controla os campos de auth da Conexão."""
        auth_type = self.auth_type_combo.get()
        state = "disabled" if auth_type == "Autenticação do Windows" else "normal"
        
        self.logon_entry.config(state=state)
        self.pass_entry.config(state=state)
        self.logon_label.config(state=state)
        self.pass_label.config(state=state)
        if state == "disabled":
            self.logon_entry.delete(0, tk.END)
            self.pass_entry.delete(0, tk.END)

    def _check_network_access(self, server: str, port: int, timeout: int = 3) -> bool:
        """Função auxiliar para checar VPN (copiada do WA.py)"""
        if server.startswith("tcp:"):
            server = server[4:]
        if ',' in server:
            server = server.split(',')[0]
            
        try:
            ip = socket.gethostbyname(server)
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                s.settimeout(timeout)
                s.connect((ip, port))
            return True
        except socket.timeout:
            print(f"DEBUG: Timeout ao tentar conectar a {server}:{port}")
            return False
        except (socket.gaierror, socket.herror):
            print(f"DEBUG: Erro de DNS ao resolver {server}")
            return False
        except Exception as e:
            if "Connection refused" in str(e):
                print(f"DEBUG: Conexão recusada por {server}:{port} (Servidor está online)")
                return True
            print(f"DEBUG: Erro de socket inesperado: {e}")
            return False

    def connect_db(self):
        """Conecta ao banco de dados (lógica V2 do WA.py)."""
        server = self.server_entry.get()
        database = self.db_entry.get()
        auth_type = self.auth_type_combo.get()
        
        if not all([server, database]):
            messagebox.showwarning("Campos Vazios", "Servidor e Banco de Dados são obrigatórios.")
            return

        self.status_label.config(text="Status: Verificando rede (VPN)...", foreground="orange")
        self.update_idletasks() 

        if not self._check_network_access(server, 1433):
            self.status_label.config(text="Status: Desconectado", foreground="red")
            messagebox.showerror("Erro de Rede", f"Não foi possível acessar o servidor: {server}.\n\nVerifique sua conexão VPN.")
            return
        
        self.status_label.config(text="Status: Conectando ao banco de dados...", foreground="orange")
        self.update_idletasks()
        
        self._close_connection()
        
        conn_string = ""
        password = "" 
        try:
            server_to_use = server
            if server.endswith(".database.windows.net") and not server.startswith('tcp:'):
                server_to_use = f"tcp:{server}"

            auth_part = ""
            if auth_type == "Autenticação do SQL Server":
                username = self.logon_entry.get()
                password = self.pass_entry.get()
                if not all([username, password]):
                    messagebox.showwarning("Campos Vazios", "Logon e Senha são obrigatórios.")
                    self.status_label.config(text="Status: Desconectado", foreground="red")
                    return
                auth_part = f"UID={username};PWD={password};"
            elif auth_type == "Autenticação do Windows":
                auth_part = "Trusted_Connection=yes;"
            else:
                raise ValueError("Tipo de autenticação inválido.")

            conn_string = (
                f"DRIVER={{ODBC Driver 17 for SQL Server}};"
                f"SERVER={server_to_use};"
                f"DATABASE={database};"
                f"{auth_part}"
                f"Encrypt=yes;"
                f"TrustServerCertificate=yes;" # Padrão V2
            )
            
            self.conn = pyodbc.connect(conn_string, timeout=5)
            self.cursor = self.conn.cursor()

            self.status_label.config(text="Status: Conectado", foreground="green")
            self.notebook.tab(1, state="normal") # Habilita Apuração
            self.notebook.select(1) # Seleciona Aba "Apuração"
            
            self._save_config()
            messagebox.showinfo("Sucesso", "Conexão estabelecida com sucesso!")

        except pyodbc.Error as e:
            self._close_connection()
            self.status_label.config(text="Status: Desconectado", foreground="red")
            messagebox.showerror("Erro de Conexão ODBC", f"O driver ODBC reportou um erro:\n\n{e}")
        except Exception as e:
            self._close_connection()
            self.status_label.config(text="Status: Desconectado", foreground="red")
            messagebox.showerror("Erro Inesperado", f"Ocorreu um erro na lógica de conexão:\n\n{e}")
            
    def start_load_data_thread(self):
        """Inicia a thread para carregar todas as 4 queries."""
        if not self.conn or not self.cursor:
            messagebox.showerror("Sem Conexão", "Por favor, conecte-se primeiro.")
            return
        
        # Limpa todas as tabelas
        self._clear_results_tree(self.tree_plano)
        self._clear_results_tree(self.tree_base)
        self._clear_results_tree(self.tree_detalhe)
        self._clear_results_tree(self.tree_credito)

        # Desabilita botões para evitar cliques duplos
        self.load_data_button.config(state="disabled")
        self.export_button.config(state="disabled")
        self.report_status_label.config(text="Iniciando carga...", foreground="blue")
        
        thread = threading.Thread(target=self._run_all_queries_thread)
        thread.daemon = True
        thread.start()

    def _run_all_queries_thread(self):
        """
        [THREAD] Executa as 4 queries em sequência.
        Cria sua própria conexão para segurança da thread.
        """
        try:
            # --- Cria uma NOVA conexão para esta thread ---
            server = self.server_entry.get()
            database = self.db_entry.get()
            auth_type = self.auth_type_combo.get()
            server_to_use = f"tcp:{server}" if server.endswith(".windows.net") and not server.startswith('tcp:') else server
            
            auth_part = ""
            if auth_type == "Autenticação do SQL Server":
                auth_part = f"UID={self.logon_entry.get()};PWD={self.pass_entry.get()};"
            elif auth_type == "Autenticação do Windows":
                auth_part = "Trusted_Connection=yes;"
            
            conn_string = (f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server_to_use};DATABASE={database};"
                           f"{auth_part}Encrypt=yes;TrustServerCertificate=yes;")
            
            with pyodbc.connect(conn_string, timeout=10) as thread_conn:
                with thread_conn.cursor() as thread_cursor:
                    
                    # --- Query 1: Plano ---
                    self.query_queue.put(("status", "Executando Query 1/4 (Planos)..."))
                    thread_cursor.execute(self.QUERY_PLANO)
                    cols = [col[0] for col in thread_cursor.description]
                    self.query_queue.put(("plano_cols", cols))
                    rows = thread_cursor.fetchall()
                    for row in rows:
                        self.query_queue.put(("plano_data", tuple(row)))

                    # --- Query 2: Base Cliente ---
                    self.query_queue.put(("status", "Executando Query 2/4 (Base Cliente)..."))
                    thread_cursor.execute(self.QUERY_BASE_CLIENTE)
                    cols = [col[0] for col in thread_cursor.description]
                    self.query_queue.put(("base_cols", cols))
                    rows = thread_cursor.fetchall()
                    for row in rows:
                        self.query_queue.put(("base_data", tuple(row)))

                    # --- Query 3: Detalhamento ---
                    self.query_queue.put(("status", "Executando Query 3/4 (Detalhamento)... (Isso pode demorar)"))
                    thread_cursor.execute(self.QUERY_DETALHAMENTO)
                    cols = [col[0] for col in thread_cursor.description]
                    self.query_queue.put(("detalhe_cols", cols))
                    rows = thread_cursor.fetchall()
                    for row in rows:
                        self.query_queue.put(("detalhe_data", tuple(row)))

                    # --- Query 4: Crédito ---
                    self.query_queue.put(("status", "Executando Query 4/4 (Crédito)..."))
                    thread_cursor.execute(self.QUERY_CREDITO)
                    cols = [col[0] for col in thread_cursor.description]
                    self.query_queue.put(("credito_cols", cols))
                    rows = thread_cursor.fetchall()
                    for row in rows:
                        self.query_queue.put(("credito_data", tuple(row)))

            self.query_queue.put(("done", "Carga de dados concluída com sucesso."))
                        
        except pyodbc.Error as e:
            self.query_queue.put(("error", f"Erro de Banco de Dados na Thread:\n{e}"))
        except Exception as e:
            self.query_queue.put(("error", f"Erro inesperado na Thread:\n{e}"))

    def _monitor_queue(self):
        """Verifica a fila por mensagens da thread e atualiza a UI."""
        try:
            while True:
                msg = self.query_queue.get_nowait()
                msg_type, msg_data = msg

                if msg_type == "status":
                    self.report_status_label.config(text=msg_data, foreground="blue")
                
                # --- Handlers Tabela 1 ---
                elif msg_type == "plano_cols":
                    self._setup_tree_columns(self.tree_plano, msg_data)
                elif msg_type == "plano_data":
                    self.tree_plano.insert(parent="", index="end", values=msg_data)
                
                # --- Handlers Tabela 2 ---
                elif msg_type == "base_cols":
                    self._setup_tree_columns(self.tree_base, msg_data)
                elif msg_type == "base_data":
                    self.tree_base.insert(parent="", index="end", values=msg_data)

                # --- Handlers Tabela 3 ---
                elif msg_type == "detalhe_cols":
                    self._setup_tree_columns(self.tree_detalhe, msg_data)
                elif msg_type == "detalhe_data":
                    self.tree_detalhe.insert(parent="", index="end", values=msg_data)

                # --- Handlers Tabela 4 ---
                elif msg_type == "credito_cols":
                    self._setup_tree_columns(self.tree_credito, msg_data)
                elif msg_type == "credito_data":
                    self.tree_credito.insert(parent="", index="end", values=msg_data)

                # --- Conclusão ---
                elif msg_type == "done":
                    self.report_status_label.config(text=msg_data, foreground="green")
                    self.load_data_button.config(state="normal")
                    self.export_button.config(state="normal")
                
                # --- Erro ---
                elif msg_type == "error":
                    self.report_status_label.config(text=msg_data, foreground="red")
                    messagebox.showerror("Erro na Carga de Dados", msg_data)
                    self.load_data_button.config(state="normal") # Reabilita para tentar de novo
        
        except queue.Empty:
            pass # Fila vazia, normal
        
        self.after(100, self._monitor_queue)

    def export_all_to_excel(self):
        """Exporta os dados das 4 tabelas para um único arquivo .xlsx"""
        if not self.tree_plano.get_children() and not self.tree_base.get_children():
            messagebox.showwarning("Nada para Exportar", "Primeiro, clique em 'Carregar Dados'.")
            return

        try:
            date_str = datetime.now().strftime('%Y%m%d')
            filename = f"ApuraçãoWebPhone_MesAtual_{date_str}.xlsx"
            
            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")],
                title="Salvar Apuração WebPhone",
                initialfile=filename
            )
            if not filepath:
                return 

            # Criar um novo Workbook
            wb = Workbook()
            # Remover a planilha padrão que é criada
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"]) 

            # Escrever cada Treeview em uma planilha separada
            self._write_tree_to_worksheet(wb, self.tree_plano, "1. Atualização de Plano")
            self._write_tree_to_worksheet(wb, self.tree_base, "2. Base de Cliente")
            self._write_tree_to_worksheet(wb, self.tree_detalhe, "3. Detalhamento")
            self._write_tree_to_worksfheet(wb, self.tree_credito, "4. Base de Crédito")

            # Salvar o arquivo
            wb.save(filepath)
                
            messagebox.showinfo("Exportação Concluída", f"Resultados salvos com sucesso em:\n{filepath}")
        except Exception as e:
            messagebox.showerror("Erro na Exportação", f"Não foi possível salvar o arquivo Excel:\n{e}")

    def _write_tree_to_worksheet(self, wb: Workbook, tree: ttk.Treeview, sheet_name: str):
        """Função auxiliar para escrever um Treeview em uma nova planilha do Workbook."""
        ws = wb.create_sheet(title=sheet_name)
        
        columns = tree['columns']
        if not columns:
            ws.cell(row=1, column=1, value="(Sem dados)")
            return
            
        # Escreve o cabeçalho
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
        
        # Escreve os dados
        for row_idx, item_id in enumerate(tree.get_children(), 2): # Começa na linha 2
            row_values = tree.item(item_id)['values']
            
            for col_idx, value in enumerate(row_values, 1):
                # Tenta converter para número se parecer um número (melhora o Excel)
                try:
                    if isinstance(value, (int, float)):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                    elif isinstance(value, str) and (value.isdigit() or re.match(r'^-?\d+\.\d+$', value)):
                         ws.cell(row=row_idx, column=col_idx, value=float(value))
                    elif isinstance(value, str) and (re.match(r'^-?\d+,\d+$', value)): # Trata números com vírgula
                         ws.cell(row=row_idx, column=col_idx, value=float(value.replace('.','').replace(',','.')))
                    else:
                        ws.cell(row=row_idx, column=col_idx, value=value)
                except:
                     ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-ajustar colunas para melhor visualização
        for col_idx, col_name in enumerate(columns, 1):
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].autosize = True

    def _clear_results_tree(self, tree_widget: ttk.Treeview):
        """Limpa todos os dados do Treeview de resultados fornecido."""
        for item in tree_widget.get_children():
            tree_widget.delete(item)
        tree_widget["columns"] = ()
        
    def _setup_tree_columns(self, tree_widget: ttk.Treeview, columns: list):
        """Configura as colunas de um Treeview."""
        tree_widget["columns"] = columns
        tree_widget.column("#0", width=0, stretch=tk.NO)
        tree_widget.heading("#0", text="", anchor="center")
        for col in columns:
            tree_widget.column(col, anchor="w", width=120)
            tree_widget.heading(col, text=col, anchor="w")

    def _close_connection(self):
        """Fecha a Conexão."""
        if self.cursor: self.cursor.close()
        if self.conn: self.conn.close()
        self.conn, self.cursor = None, None
        self.status_label.config(text="Status: Desconectado", foreground="red")
        self.notebook.tab(1, state="disabled") # Desabilita Apuração
        self.export_button.config(state="disabled") # Desabilita exportar
        
    def _save_config(self):
        """Salva as configurações de conexão no arquivo .ini"""
        try:
            if not self.config.has_section('Connection'):
                self.config.add_section('Connection')
            self.config.set('Connection', 'server', self.server_entry.get())
            self.config.set('Connection', 'database', self.db_entry.get())
            self.config.set('Connection', 'logon', self.logon_entry.get())
            self.config.set('Connection', 'password', self.pass_entry.get()) 

            with open(CONFIG_FILE, 'w') as configfile:
                self.config.write(configfile)
        except Exception as e:
            print(f"Erro ao salvar config: {e}")

    def _load_config(self):
        """Carrega as configurações do arquivo .ini ao iniciar"""
        if not os.path.exists(CONFIG_FILE):
            return
            
        try:
            self.config.read(CONFIG_FILE)
            
            if 'Connection' in self.config:
                cfg = self.config['Connection']
                self.server_entry.delete(0, tk.END)
                self.server_entry.insert(0, cfg.get('server', '')) # <-- CORRIGIDO
                self.db_entry.delete(0, tk.END)
                self.db_entry.insert(0, cfg.get('database', '')) # <-- CORRIGIDO
                self.logon_entry.delete(0, tk.END)
                self.logon_entry.insert(0, cfg.get('logon', '')) # <-- CORRIGIDO
                self.pass_entry.delete(0, tk.END)
                self.pass_entry.insert(0, cfg.get('password', ''))
                
        except Exception as e:
            print(f"Erro ao carregar config: {e}")

    def on_closing(self):
        """Chamado quando a janela é fechada."""
        if messagebox.askokcancel("Sair", "Deseja fechar a aplicação?"):
            self._close_connection()
            self.destroy()

if __name__ == "__main__":
    try:
        import pyodbc
    except ImportError:
        print("Erro: A biblioteca 'pyodbc' não está instalada.")
        print("Por favor, instale-a usando o comando: pip install pyodbc")
        exit()
        
    try:
        import openpyxl
    except ImportError:
        print("Erro: A biblioteca 'openpyxl' não está instalada.")
        print("Por favor, instale-a usando o comando: pip install openpyxl")
        exit()

    app = WebPhoneReportTool()
    app.protocol("WM_DELETE_WINDOW", app.on_closing)
    app.mainloop()
